import csv
from pathlib import Path
import fitz
from docx import Document
from openpyxl import load_workbook


PLAIN_TEXT_SUFFIXES = {
    ".txt",
    ".md",
    ".py",
    ".js",
    ".ts",
    ".java",
    ".json",
    ".yaml",
    ".yml",
    ".xml",
    ".html",
    ".css",
    ".log",
}


def load_txt(path: str) -> str:
    with open(path, "r", encoding="utf-8") as f:
        return f.read()


def load_pdf(path: str) -> str:
    doc = fitz.open(path)

    text = ""

    try:
        for page in doc:
            text += page.get_text()

    finally:
        doc.close()

    return text


def load_docx(path: str) -> str:
    document = Document(path)
    parts = []

    for block in document.iter_inner_content():
        if hasattr(block, "rows"):
            for row in block.rows:
                cells = [
                    cell.text.strip()
                    for cell in row.cells
                ]

                if any(cells):
                    parts.append(" | ".join(cells))

        elif block.text.strip():
            parts.append(block.text)

    return "\n".join(parts)


def load_xlsx(path: str) -> str:
    workbook = load_workbook(
        path,
        data_only=True,
        read_only=True
    )

    parts = []

    try:
        for sheet in workbook.worksheets:
            parts.append(f"Sheet: {sheet.title}")

            for row in sheet.iter_rows(values_only=True):
                values = [
                    "" if value is None else str(value)
                    for value in row
                ]

                if any(value.strip() for value in values):
                    parts.append(" | ".join(values))

    finally:
        workbook.close()

    return "\n".join(parts)


def load_csv(path: str) -> str:
    parts = []

    with open(path, "r", encoding="utf-8-sig", newline="") as f:
        reader = csv.reader(f)

        for row in reader:
            values = [
                value.strip()
                for value in row
            ]

            if any(values):
                parts.append(" | ".join(values))

    return "\n".join(parts)


def load_document(path: str) -> str:
    suffix = Path(path).suffix.lower()

    if suffix in PLAIN_TEXT_SUFFIXES:
        return load_txt(path)

    elif suffix == ".pdf":
        return load_pdf(path)

    elif suffix == ".docx":
        return load_docx(path)

    elif suffix == ".xlsx":
        return load_xlsx(path)

    elif suffix == ".csv":
        return load_csv(path)

    else:
        raise ValueError(f"Unsupported file type: {suffix}")
