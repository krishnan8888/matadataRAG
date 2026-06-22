import csv
import json
import re
import shutil
from pathlib import Path

from docx import Document
from openpyxl import load_workbook


PROJECT_ROOT = Path(__file__).resolve().parents[2]
APP_DIR = PROJECT_ROOT / "app"

DATAFRAME_DIR = APP_DIR / "dataframes"
TABLE_DIR = APP_DIR / "tables"
STRUCTURED_JSON_DIR = APP_DIR / "structured"
KEYWORD_INDEX_DIR = APP_DIR / "keyword_index"
RAW_FILE_DIR = APP_DIR / "raw_files"


def write_json(path: Path, data) -> Path:
    path.parent.mkdir(
        parents=True,
        exist_ok=True
    )

    with open(path, "w", encoding="utf-8") as f:
        json.dump(data, f, indent=4)

    return path


def read_csv_rows(file_path: Path) -> list[dict]:
    with open(file_path, "r", encoding="utf-8-sig", newline="") as f:
        reader = csv.DictReader(f)
        rows = []

        for row in reader:
            rows.append(dict(row))

    return rows


def read_xlsx_sheets(file_path: Path) -> list[dict]:
    workbook = load_workbook(
        file_path,
        data_only=True,
        read_only=True
    )

    sheets = []

    try:
        for sheet in workbook.worksheets:
            rows = list(sheet.iter_rows(values_only=True))

            if not rows:
                sheets.append({
                    "sheet_name": sheet.title,
                    "columns": [],
                    "rows": []
                })
                continue

            headers = [
                f"column_{index + 1}" if value is None else str(value)
                for index, value in enumerate(rows[0])
            ]

            records = []

            for row in rows[1:]:
                record = {}

                for index, header in enumerate(headers):
                    value = row[index] if index < len(row) else None
                    record[header] = value

                if any(value is not None for value in record.values()):
                    records.append(record)

            sheets.append({
                "sheet_name": sheet.title,
                "columns": headers,
                "rows": records
            })

    finally:
        workbook.close()

    return sheets


def extract_docx_tables(file_path: Path) -> list[dict]:
    document = Document(file_path)
    tables = []
    current_document_section = ""
    current_project = ""
    current_section = ""
    table_index = 0

    for block in document.iter_inner_content():
        if not hasattr(block, "rows"):
            text = block.text.strip()

            if not text:
                continue

            if (
                text.startswith("Statement of Work")
                or text.startswith("Change Order")
            ):
                current_document_section = text

            project_match = re.search(
                r'Project\s+"?([A-Za-z0-9_-]+)',
                text,
                flags=re.IGNORECASE,
            )

            if project_match:
                current_project = project_match.group(1).strip()

            if re.match(r"^\d+\.\s+", text):
                current_section = text

            continue

        table = block
        rows = []

        for row in table.rows:
            rows.append([
                cell.text.strip()
                for cell in row.cells
            ])

        tables.append({
            "table_id": f"{file_path.stem}_table_{table_index}",
            "source_file": file_path.name,
            "source_type": "docx",
            "document_section": current_document_section,
            "project_id": current_project,
            "section_title": current_section,
            "rows": rows
        })
        table_index += 1

    return tables


def extract_xlsx_tables(file_path: Path) -> list[dict]:
    tables = []

    for sheet in read_xlsx_sheets(file_path):
        tables.append({
            "table_id": f"{file_path.stem}_{sheet['sheet_name']}",
            "source_file": file_path.name,
            "source_type": "xlsx",
            "caption": sheet["sheet_name"],
            "columns": sheet["columns"],
            "rows": sheet["rows"]
        })

    return tables


def extract_csv_table(file_path: Path) -> list[dict]:
    rows = read_csv_rows(file_path)
    columns = list(rows[0].keys()) if rows else []

    return [{
        "table_id": file_path.stem,
        "source_file": file_path.name,
        "source_type": "csv",
        "columns": columns,
        "rows": rows
    }]


def store_dataframe(document_id: str, file_path: Path) -> Path | None:
    suffix = file_path.suffix.lower()

    if suffix == ".xlsx":
        data = {
            "document_id": document_id,
            "source_file": str(file_path),
            "sheets": read_xlsx_sheets(file_path)
        }

    elif suffix == ".csv":
        data = {
            "document_id": document_id,
            "source_file": str(file_path),
            "rows": read_csv_rows(file_path)
        }

    else:
        return None

    return write_json(DATAFRAME_DIR / f"{document_id}.json", data)


def store_tables(document_id: str, file_path: Path) -> Path | None:
    suffix = file_path.suffix.lower()

    if suffix == ".docx":
        tables = extract_docx_tables(file_path)

    elif suffix == ".xlsx":
        tables = extract_xlsx_tables(file_path)

    elif suffix == ".csv":
        tables = extract_csv_table(file_path)

    else:
        tables = []

    if not tables:
        return None

    return write_json(
        TABLE_DIR / f"{document_id}.json",
        {
            "document_id": document_id,
            "source_file": str(file_path),
            "tables": tables
        }
    )


def store_structured_json(document_id: str, metadata: dict) -> Path:
    data = {
        "document_id": document_id,
        "document_type": metadata.get("document_type"),
        "summary": metadata.get("summary"),
        "topics": metadata.get("topics", []),
        "important_keywords": metadata.get("important_keywords", []),
        "possible_user_queries": metadata.get("possible_user_queries", [])
    }

    return write_json(STRUCTURED_JSON_DIR / f"{document_id}.json", data)


def store_keyword_index(document_id: str, text: str, metadata: dict) -> Path:
    terms = set()

    for keyword in metadata.get("important_keywords", []):
        terms.add(keyword.lower())

    for topic in metadata.get("topics", []):
        terms.add(topic.lower())

    for token in text.replace("|", " ").split():
        cleaned = token.strip(".,:;()[]{}\"'").lower()

        if len(cleaned) >= 3:
            terms.add(cleaned)

    data = {
        "document_id": document_id,
        "terms": sorted(terms)
    }

    return write_json(KEYWORD_INDEX_DIR / f"{document_id}.json", data)


def store_raw_file(document_id: str, file_path: Path) -> Path:
    RAW_FILE_DIR.mkdir(
        parents=True,
        exist_ok=True
    )

    target = RAW_FILE_DIR / f"{document_id}{file_path.suffix}"
    shutil.copy2(file_path, target)

    return target
